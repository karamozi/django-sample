from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from accounts.models import Schools, Classes, Table
from openpyxl import Workbook, load_workbook
import os
from django.conf import settings
from django.contrib.auth.decorators import login_required
from  django.core import validators
from django.core.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from django.views.generic.edit import UpdateView

# Create your views here.

# from django.contrib.auth import get_user_model
# current_User = get_user_model()


filename = os.path.join(settings.BASE_DIR, 'static', 'Three_Schools.xlsx')
wb = load_workbook(filename)
ws = wb['Data']

schools = list()

for i in range(1, 156):

    if ws['r'+str(i)].value:
        schools.append(ws['r'+str(i)].value)

# print(schools)


def userCreate(request):

    email_message = str()
    username_message = str()
    password_message = str()

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        emails = [i.email for i in User.objects.all()]
        usernames = [i.username for i in User.objects.all()]


        if not email:
            email_message = 'وارد کردن ایمیل الزامی است'
        elif email in emails:
            email_message = 'این ایمیل قبلا استفاده شده است'
        else:
            try:
                validators.validate_email(email)
            except ValidationError:
                email_message = 'ایمیل اشتباه است'

        if not username:
            username_message = 'نام کاربری الزامی است'
        elif username in usernames:
            username_message = 'این نام قبلا استفاده شده است'

        if not password:
            password_message = 'رمز عبور را وارد کنید'
        elif password != confirm_password:
            password_message = 'رمز عبور یکسان نیست'


        if email_message == '' and username_message == '' and password_message == '':
            user = User.objects.create_user(username,email,password)
            user.save()

            return redirect(reverse('login'))


    return render(request,'signup.html',{'email_message':email_message,
                                         'username_message':username_message,
                                         'password_message':password_message})

@login_required
def school_selection(request):
    if request.method == 'POST':
        valueList = request.POST.get('valueList')
        item = Schools.objects.filter(user=request.user).all().delete()
        if valueList:
            valueList = valueList.split(',')

            for i in valueList:
                new_item = Schools(school=i,user=request.user)
                new_item.save()


    selected_schools = [i.school for i in Schools.objects.filter(user=request.user)]
    remaining_schools = [i for i in schools if i not in selected_schools]

    return render(request,'school_selection.html',{'remaining_schools':remaining_schools,
                                                  'selected_schools':selected_schools})

@login_required
def class_selection(request):
    school = ''
    all = True
    if request.method == 'POST':
        school = request.POST.get('school')
        capasity = request.POST.get('capasity')



        item = Classes(school=school, capasity=capasity, user=request.user)
        item.save()
        all = False

    class_list = Classes.objects.filter(user=request.user)
    selected_schools = [i.school for i in Schools.objects.filter(user=request.user)]



    return render(request,'class_selection.html',{'selected_schools':selected_schools,
                                                  'class_list':class_list,
                                                  'last_school':school,
                                                  'all':all})


@login_required
def table(request):

    if request.method == 'POST':

        school = request.POST.get("school")
        course = request.POST.get("course")
        num_of_session = request.POST.get("num_of_session")
        teacher = request.POST.get("teacher")
        first_day = request.POST.get("first_day")
        if num_of_session == 'دو روز در هفته':
            second_day = request.POST.get("second_day")
        else:
            second_day = ""
        start_time = request.POST.get("start_time")
        end_time = request.POST.get("end_time")
        signup_capasity = request.POST.get("signup_capasity")

        item = Table(user=request.user,school=school,course=course,
                    num_of_session=num_of_session,teacher=teacher,
                    first_day=first_day,second_day=second_day,
                    start_time=start_time,end_time=end_time,signup_capasity=signup_capasity)

        item.save()

    table_list = Table.objects.filter(user=request.user)
    selected_schools = [i.school for i in Schools.objects.filter(user=request.user)]

    hour_list = range(6,20)
    minute_list = ['00','15','30','45']

    time_list = list()
    for hour in hour_list:
        for minute in minute_list:
            time = str(hour)+':'+str(minute)
            time_list.append(time)


    return render(request,'table.html',{'selected_schools':selected_schools,
                                        'table_list':table_list,
                                        'time_list':time_list,})

def create_excel(request):

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename='+ request.user.username +'.xlsx'
    class_list = Classes.objects.filter(user=request.user)
    table_list = Table.objects.filter(user=request.user)

    # print(unique, len(unique))

    unique = list(dict.fromkeys([i.school for i in class_list]))


    for i in unique:
        filter_list = list(class_list.filter(school=i))
        school_filter_list = [z for z in filter_list]
        for j in filter_list:

            print((unique.index(j.school)+1)*100 + filter_list.index(j)+1, j.school)


    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws2 = wb.create_sheet("Capasity")


    columns = [('a1','ردیف'),('b1','دانشکده'),('c1','نام درس'),('d1','تعداد جلسات در هفته'),('e1','نام استاد'),('f1','اولین روز درس'),
               ('g1','ساعت شروع'),('h1','ساعت پایان'),('i1','دومین روز درس'),('j1','ساعت شروع'),('k1','ساعت پایان'),('l1','ظرفیت ثبت نامی احتمالی درس')]

    for i in columns:
        ws1[i[0]].value = i[1]
        ws1[i[0]].fill = PatternFill("solid", fgColor='E1C39B')



    ws2['a1'].value = 'شماره کلاس'
    ws2['b1'].value = 'نام دانشکده'
    ws2['c1'].value = 'ظرفیت کلاس'

    row = 2
    for object in table_list:
        ws1['a'+str(row)].value = row - 1
        ws1['b'+str(row)].value = object.school
        ws1['c'+str(row)].value = object.course
        ws1['d'+str(row)].value = object.num_of_session
        ws1['e'+str(row)].value = object.teacher
        ws1['f'+str(row)].value = object.first_day
        ws1['g'+str(row)].value = '"' + object.start_time + '"'
        ws1['h'+str(row)].value = '"' + object.end_time + '"'
        if object.num_of_session == 'دو روز در هفته':
            ws1['i'+str(row)].value = object.second_day
            ws1['j'+str(row)].value = '"' + object.start_time + '"'
            ws1['k'+str(row)].value = '"' + object.end_time + '"'
        ws1['l'+str(row)].value = str(object.signup_capasity)
        row = row+ 1

    row = 2

    for i in unique:
        filter_list = list(class_list.filter(school=i))
        school_filter_list = [z for z in filter_list]
        for j in filter_list:
            index = (unique.index(j.school)+1)*100 + filter_list.index(j)+1
            ws2['a'+str(row)].value = index
            ws2['b'+str(row)].value = j.school
            ws2['c'+str(row)].value = j.capasity
            row = row+ 1

    wb.save(response)
    return response


def delete_table(request,pk):
    item = get_object_or_404(Table, pk=pk)
    item.delete()

    return redirect ('table')

def delete_class(request,pk):
    item = get_object_or_404(Classes, pk=pk)
    item.delete()

    return redirect ('class_selection')

def updateClass(request,pk):
    item = get_object_or_404(Classes, pk=pk)
    item.capasity = request.POST.get('update_capasity')

    item.save()
    return redirect('class_selection')


#
def updateTable(request,pk):
    item = get_object_or_404(Table, pk=pk)

    item.course = request.POST.get('course')
    item.num_of_session = request.POST.get('num_of_session')
    num_of_session = request.POST.get('num_of_session')
    item.teacher = request.POST.get('teacher')
    item.first_day = request.POST.get('first_day')
    if num_of_session == 'دو روز در هفته':
        item.second_day = request.POST.get('second_day')
    else:
        item.second_day = ''
    item.start_time = request.POST.get('start_time')
    item.end_time = request.POST.get('end_time')
    item.signup_capasity = request.POST.get('signup_capasity')

    item.save()

    return redirect('table')
